import os


def add_row_to_csv(row, fp, i=0):
    """
    :param row: list
    :param fp: str
    :param i: int
    :return: row_list
    """
    data_row = row
    path = fp
    if not check_file_exist(fp, 'csv'):
        print("数据添加失败:无该文件", )
        return None
    import csv
    with open(path, 'a+', encoding="utf-8", newline="") as f:
        csv_write = csv.writer(f)
        csv_write.writerow(data_row)
    print("数据添加成功", str(row), " to ", str(fp))


def read_latest_row_from_csv(fp):
    path = fp
    if not check_file_exist(fp, 'csv'):
        return None
    import csv
    data_row = []
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for i in range(len(lines) - 1, 0, -1):
            if lines[i]:
                return lines[i].split(",")


def check_file_exist(f, etn):
    print("检查文件是否存在：", f, end="--")
    if not os.path.exists(f):
        print("No such file")
        return False
    if not os.path.splitext(f)[-1] == "." + etn:
        print(os.path.splitext(f)[-1] + "  file type is not ." + etn)
        return False
    print("存在！")
    return True


def read_excel_to_row_list(fp, i=0):
    """
    :param fp: str
    :param i: int
    :return: row_list
    """

    check_file_exist(fp, 'xlsx')
    import openpyxl
    row_list = []
    wb = openpyxl.load_workbook(fp)

    ws = wb[wb.sheetnames[i]]
    for row in ws.rows:
        row_value_list = []
        for cell in row:
            row_value_list.append(cell.value)
        row_list.append(row_value_list)
    return row_list


def read_csv(fp):
    if not check_file_exist(fp, 'csv'):
        return None

    import csv
    if isinstance(fp, str):
        data_list = []
        with open(fp, 'r', encoding='utf_8_sig') as f:
            csv_file = csv.reader(f)
            for row in csv_file:  # 从CSV里挑选数据
                data_list.append(row)
    return data_list


def xlsx_to_orderedDict_by_Panda(fp):
    if not check_file_exist(fp, 'xlsx') or not check_file_exist(fp, 'xls'):
        return None

    import pandas as pd
    csv_fp_list = []
    print("load file:", fp)
    data_ordered_dict = pd.read_excel(fp, sheet_name=None)
    print(fp + "converted to orderedDict Successfully ")
    return data_ordered_dict


def get_classify_dict_from_excel(fp, i=0):
    """
    :param fp: str
    :param i: int
    :return: classify_dict
    """
    if not check_file_exist(fp, 'xlsx') or not check_file_exist(fp, 'xls'):
        return None
    import openpyxl
    classify_dict = {}
    wb = openpyxl.load_workbook(fp)
    sn = wb.sheetnames[i]
    ws = get_sheet_by_name(sn)
    print("获取对照 %s :%s-%s" % (sn, ws.rows[0][0], ws.rows[0][1]))
    for row in ws.rows:
        classify_dict[row[0]] = row[1]
    return classify_dict


if __name__ == '__main__':
    fp = r'D:\GitHub\freightQuery\freightQuery\freightdaily.csv'
    latest_row = read_latest_row_from_csv(fp)
    print(latest_row)
